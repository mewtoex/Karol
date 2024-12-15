import os
from tkinter import Tk, Button, filedialog, Label
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def processar_xlsx(arquivo):
    wb = load_workbook(arquivo)
    ws = wb.active

    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]  
        text = cell.value
        if text:
            parts = text.split('-')
            arrays = [part.split() for part in parts] 
            print(arrays)

            valid_arrays = [arr for arr in arrays if len(arr) == 3]

            if len(valid_arrays) > 1:
                for i in range(1, len(valid_arrays)):
                    prev_array = valid_arrays[i - 1]
                    current_array = valid_arrays[i]
                    print(valid_arrays)



                    if prev_array[1] != current_array[1] and prev_array[2] != current_array[2]:
                        for cell_to_color in row:
                            cell_to_color.fill = fill_red
                        break

    nome_arquivo, extensao = os.path.splitext(arquivo)
    novo_arquivo = f"{nome_arquivo}_processado{extensao}"
    wb.save(novo_arquivo)
    return novo_arquivo

def selecionar_arquivo():
    caminho_arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if caminho_arquivo:
        arquivo_processado = processar_xlsx(caminho_arquivo)
        label_resultado.config(text=f"Arquivo salvo: {arquivo_processado}")

janela = Tk()
janela.title("Processador de Arquivo XLSX")

botao_selecionar = Button(janela, text="Selecionar Arquivo XLSX", command=selecionar_arquivo)
botao_selecionar.pack(pady=10)

label_resultado = Label(janela, text="")
label_resultado.pack(pady=10)

janela.mainloop()